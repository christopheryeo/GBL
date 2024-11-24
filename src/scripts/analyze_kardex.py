import pandas as pd
import openpyxl

# Load the workbook to get sheet names
wb = openpyxl.load_workbook('/Users/chrisyeo/Library/CloudStorage/OneDrive-Personal/Dev/windsurf/GBL/uploads/Kardex_for_vehicle_6_years_old.xlsx', data_only=True)
print("\nSheet names:", wb.sheetnames)

# Read the Excel file
for sheet_name in wb.sheetnames:
    print(f"\nAnalyzing sheet: {sheet_name}")
    df = pd.read_excel('/Users/chrisyeo/Library/CloudStorage/OneDrive-Personal/Dev/windsurf/GBL/uploads/Kardex_for_vehicle_6_years_old.xlsx', 
                      sheet_name=sheet_name,
                      header=3)  # Headers start from row 4 (0-based index 3)
    
    print("\nColumns found:")
    for col in df.columns:
        print(f"- {col}")
